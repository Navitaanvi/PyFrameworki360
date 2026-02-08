from openpyxl import load_workbook
import os

def read_excel():
    base_path = os.path.dirname(os.path.dirname(__file__))
    excel_path = os.path.join(base_path, "resources", "testdata.xlsx")

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "username": row[0],
            "password": row[1],
            "test_case": row[2]
        })

    return data
